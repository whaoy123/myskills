#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""Extract purchased solderable electronic components from invoice data.

The workflow is intentionally conservative:
1. normalize source rows from PDF/XLSX/CSV;
2. classify every row as include/exclude/review;
3. aggregate only records that are safe to merge;
4. write traceable CSV/XLSX outputs and a validation report.

This script describes what appears in procurement evidence. It does not decide
what a PCB actually requires; BOM/netlist/schematic reconciliation belongs to
the downstream soldering-table workflow.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import defaultdict
from pathlib import Path
from typing import Any

import pdfplumber
import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


FEE_KEYWORDS = ["运费", "配送费", "包装费", "快递费", "服务费", "优惠", "折扣"]
NON_COMPONENT_KEYWORDS = [
    "手套", "劳保", "工具", "钳子", "镊子", "吸锡器", "线路板", "印制电路板", "pcb裸板",
    "打板", "塑壳", "外壳", "外罩", "外壳组件", "螺丝", "螺母", "铜柱", "扎带", "导热胶",
    "胶水", "焊锡", "吸锡带", "静电环", "周转箱", "标签纸",
]
COMPONENT_KEYWORDS = [
    "电阻", "电容", "电感", "芯片", "集成电路", "传感器", "二极管", "三极管", "mosfet",
    "场效应管", "连接器", "端子", "晶振", "继电器", "开关", "插针", "排母", "排针", "针座",
    "运放", "放大器", "隔离器", "稳压器", "基准", "保险丝", "光耦", "变压器", "电源模块",
]

HEADER_ALIASES = {
    "name": ["项目名称", "商品名称", "货物或应税劳务名称", "货物名称", "名称", "品名"],
    "model": ["规格型号", "型号", "规格", "型号规格"],
    "qty": ["数量", "采购数量", "数 量"],
    "unit": ["单位", "单 位"],
}

PACKAGE_PATTERNS = [
    (r"\b(0201|0402|0603|0805|1206|1210|1812|2512)\b", lambda m: f"{m.group(1)} (SMD)"),
    (r"\b(SOIC[- ]?\d+|SOP[- ]?\d+|TSSOP[- ]?\d+|SSOP[- ]?\d+)\b", lambda m: f"{m.group(1).upper()} (SMD)"),
    (r"\b(QFN[- ]?\d+|QFP[- ]?\d+|LQFP[- ]?\d+)\b", lambda m: f"{m.group(1).upper()} (SMD)"),
    (r"\b(SOT[- ]?\d+[A-Z]?)\b", lambda m: f"{m.group(1).upper()} (SMD)"),
]


def normalize_space(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def to_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    text = normalize_space(value).replace(",", "")
    try:
        return float(text)
    except ValueError:
        return None


def location_label(record: dict[str, Any]) -> str:
    if record.get("source_page"):
        return f"page {record['source_page']} line {record.get('source_row', '')}".strip()
    if record.get("source_sheet"):
        return f"sheet {record['source_sheet']} row {record.get('source_row', '')}".strip()
    return f"row {record.get('source_row', '')}".strip()


def classify_decision(name: str, model: str = "", raw_text: str = "", qty: float | None = None) -> tuple[str, str]:
    text = normalize_space(f"{name} {model} {raw_text}").lower()

    if qty is not None and qty < 0:
        return "exclude", "negative_quantity_or_discount"
    if re.search(r"(?:^|\s)-\d+(?:\.\d+)?(?:\s|$)", raw_text):
        return "exclude", "negative_amount_or_discount"
    for keyword in FEE_KEYWORDS:
        if keyword.lower() in text:
            return "exclude", f"fee_or_discount:{keyword}"
    for keyword in NON_COMPONENT_KEYWORDS:
        if keyword.lower() in text:
            return "exclude", f"non_component:{keyword}"
    for keyword in COMPONENT_KEYWORDS:
        if keyword.lower() in text:
            return "include", f"component_keyword:{keyword}"
    if re.search(r"\b(?:IC|MOS|LED|TVS|D-SUB|DB\d+|DR\d+|RJ\d+|BNC|SMA)\b", text, re.I):
        return "include", "component_pattern"
    return "review", "insufficient_evidence"


def infer_category(name: str, model: str = "") -> str:
    text = normalize_space(f"{name} {model}").lower()
    if "电容" in text:
        return "电容"
    if "电阻" in text:
        return "电阻"
    if "电感" in text:
        return "电感"
    if any(key in text for key in ["端子", "接线"]):
        return "接线端子"
    if any(key in text for key in ["连接器", "d-sub", "插针", "排针", "排母", "针座"]) or re.search(r"\b(?:DB|DR)\d+\b", text, re.I):
        return "连接器"
    if "继电器" in text:
        return "继电器"
    if "晶振" in text:
        return "晶振"
    if "保险丝" in text:
        return "保险丝"
    if any(key in text for key in ["芯片", "集成电路", "传感器", "运放", "放大器", "隔离器", "稳压器", "光耦", "mosfet", "场效应管"]):
        return "芯片/模块"
    return "其他元器件"


def infer_package(name: str, model: str = "") -> str:
    text = normalize_space(f"{name} {model}")
    for pattern, formatter in PACKAGE_PATTERNS:
        match = re.search(pattern, text, re.I)
        if match:
            return formatter(match)
    pitch = re.search(r"(\d+(?:\.\d+)?)\s*mm", text, re.I)
    if pitch and any(key in text for key in ["端子", "接线", "连接器"]):
        return f"间距 {pitch.group(1)} mm (THT)"
    if any(key in text for key in ["插件", "直插", "tht", "axial"]):
        return "THT"
    if any(key in text.lower() for key in ["贴片", "smd", "smt"]):
        return "SMD"
    if any(key in text.lower() for key in ["焊线", "线端"]):
        return "线端/焊线"
    return "待确认"


def clean_model(name: str, model: str) -> str:
    candidate = normalize_space(model)
    if candidate:
        return re.sub(r"^\*[^*]+\*", "", candidate).strip()
    return re.sub(r"^\*[^*]+\*", "", normalize_space(name)).strip()


def finalize_record(record: dict[str, Any]) -> dict[str, Any]:
    decision, reason = classify_decision(
        record.get("name", ""),
        record.get("model", ""),
        record.get("raw_text", ""),
        record.get("qty"),
    )
    record["decision"] = decision
    record["reason"] = reason
    record["category"] = infer_category(record.get("name", ""), record.get("model", "")) if decision == "include" else ""
    record["package"] = infer_package(record.get("name", ""), record.get("model", "")) if decision == "include" else ""
    record["normalized_model"] = clean_model(record.get("name", ""), record.get("model", "")) if decision == "include" else ""
    return record


def parse_pdf_line(line: str, source_file: str, page_no: int, line_no: int) -> dict[str, Any] | None:
    raw = normalize_space(line)
    if not raw.startswith("*"):
        return None

    parts = raw.split()
    base = {
        "source_file": source_file,
        "source_type": "pdf",
        "source_page": page_no,
        "source_sheet": "",
        "source_row": line_no,
        "raw_text": raw,
        "name": "",
        "model": "",
        "qty": None,
        "unit": "",
    }

    if len(parts) < 6:
        base["name"] = raw
        base["decision"] = "review"
        base["reason"] = "pdf_line_parse_failed"
        base["category"] = ""
        base["package"] = ""
        base["normalized_model"] = ""
        return base

    qty = to_number(parts[-5])
    base["unit"] = parts[-6]
    base["qty"] = qty
    base["name"] = parts[0]
    base["model"] = " ".join(parts[1:-6]) if len(parts) > 7 else ""

    if qty is None and not any(keyword in raw for keyword in FEE_KEYWORDS):
        base["decision"] = "review"
        base["reason"] = "pdf_quantity_parse_failed"
        base["category"] = ""
        base["package"] = ""
        base["normalized_model"] = ""
        return base
    return finalize_record(base)


def extract_pdf(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    records: list[dict[str, Any]] = []
    warnings: list[str] = []
    try:
        with pdfplumber.open(path) as pdf:
            for page_idx, page in enumerate(pdf.pages, 1):
                text = page.extract_text() or ""
                if not text.strip():
                    warnings.append(f"{path.name}: page {page_idx} has no extractable text")
                    continue
                for line_idx, line in enumerate(text.splitlines(), 1):
                    record = parse_pdf_line(line, path.name, page_idx, line_idx)
                    if record:
                        records.append(record)
    except Exception as exc:
        warnings.append(f"{path.name}: PDF read failed: {exc}")
    return records, warnings


def find_header_index(headers: list[str], aliases: list[str]) -> int | None:
    normalized = [normalize_space(value) for value in headers]
    for idx, header in enumerate(normalized):
        if any(alias in header for alias in aliases):
            return idx
    return None


def extract_xlsx(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    records: list[dict[str, Any]] = []
    warnings: list[str] = []
    try:
        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        return [], [f"{path.name}: XLSX read failed: {exc}"]

    for sheet in workbook.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        header_row_idx = None
        column_map: dict[str, int | None] = {}
        for idx, row in enumerate(rows[:20]):
            headers = [normalize_space(value) for value in row]
            name_idx = find_header_index(headers, HEADER_ALIASES["name"])
            qty_idx = find_header_index(headers, HEADER_ALIASES["qty"])
            if name_idx is not None and qty_idx is not None:
                header_row_idx = idx
                column_map = {
                    "name": name_idx,
                    "model": find_header_index(headers, HEADER_ALIASES["model"]),
                    "qty": qty_idx,
                    "unit": find_header_index(headers, HEADER_ALIASES["unit"]),
                }
                break
        if header_row_idx is None:
            warnings.append(f"{path.name}/{sheet.title}: no recognizable invoice header")
            continue

        for row_idx, row in enumerate(rows[header_row_idx + 1 :], header_row_idx + 2):
            name = normalize_space(row[column_map["name"]]) if column_map["name"] is not None and column_map["name"] < len(row) else ""
            if not name:
                continue
            model_idx = column_map.get("model")
            unit_idx = column_map.get("unit")
            qty_idx = column_map.get("qty")
            model = normalize_space(row[model_idx]) if model_idx is not None and model_idx < len(row) else ""
            unit = normalize_space(row[unit_idx]) if unit_idx is not None and unit_idx < len(row) else ""
            qty = to_number(row[qty_idx]) if qty_idx is not None and qty_idx < len(row) else None
            raw = " | ".join(normalize_space(value) for value in row if value not in (None, ""))
            record = {
                "source_file": path.name,
                "source_type": "xlsx",
                "source_page": "",
                "source_sheet": sheet.title,
                "source_row": row_idx,
                "raw_text": raw,
                "name": name,
                "model": model,
                "qty": qty,
                "unit": unit,
            }
            if qty is None:
                record.update({"decision": "review", "reason": "xlsx_quantity_parse_failed", "category": "", "package": "", "normalized_model": ""})
            else:
                finalize_record(record)
            records.append(record)
    workbook.close()
    return records, warnings


def resolve_csv_columns(fieldnames: list[str]) -> dict[str, str | None]:
    result: dict[str, str | None] = {}
    for key, aliases in HEADER_ALIASES.items():
        result[key] = None
        for field in fieldnames:
            if any(alias in normalize_space(field) for alias in aliases):
                result[key] = field
                break
    return result


def extract_csv(path: Path) -> tuple[list[dict[str, Any]], list[str]]:
    records: list[dict[str, Any]] = []
    warnings: list[str] = []
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as handle:
            reader = csv.DictReader(handle)
            fieldnames = reader.fieldnames or []
            columns = resolve_csv_columns(fieldnames)
            if columns["name"] is None or columns["qty"] is None:
                return [], [f"{path.name}: CSV requires recognizable name and quantity columns"]
            for row_idx, row in enumerate(reader, 2):
                name = normalize_space(row.get(columns["name"] or "", ""))
                if not name:
                    continue
                model = normalize_space(row.get(columns["model"] or "", "")) if columns["model"] else ""
                unit = normalize_space(row.get(columns["unit"] or "", "")) if columns["unit"] else ""
                qty = to_number(row.get(columns["qty"] or "", ""))
                raw = " | ".join(f"{key}={normalize_space(value)}" for key, value in row.items())
                record = {
                    "source_file": path.name,
                    "source_type": "csv",
                    "source_page": "",
                    "source_sheet": "",
                    "source_row": row_idx,
                    "raw_text": raw,
                    "name": name,
                    "model": model,
                    "qty": qty,
                    "unit": unit,
                }
                if qty is None:
                    record.update({"decision": "review", "reason": "csv_quantity_parse_failed", "category": "", "package": "", "normalized_model": ""})
                else:
                    finalize_record(record)
                records.append(record)
    except Exception as exc:
        warnings.append(f"{path.name}: CSV read failed: {exc}")
    return records, warnings


def collect_input_files(input_path: Path) -> list[Path]:
    if input_path.is_file():
        return [input_path]
    if not input_path.is_dir():
        raise FileNotFoundError(input_path)
    files = []
    for suffix in ("*.pdf", "*.xlsx", "*.csv"):
        files.extend(input_path.glob(suffix))
    return sorted(set(files), key=lambda path: path.name.lower())


def extract_records(input_path: Path) -> tuple[list[dict[str, Any]], list[str], list[str]]:
    records: list[dict[str, Any]] = []
    warnings: list[str] = []
    processed: list[str] = []
    for path in collect_input_files(input_path):
        suffix = path.suffix.lower()
        if suffix == ".pdf":
            new_records, new_warnings = extract_pdf(path)
        elif suffix == ".xlsx":
            new_records, new_warnings = extract_xlsx(path)
        elif suffix == ".csv":
            new_records, new_warnings = extract_csv(path)
        else:
            warnings.append(f"{path.name}: unsupported source type")
            continue
        processed.append(path.name)
        records.extend(new_records)
        warnings.extend(new_warnings)
    return records, warnings, processed


def aggregate_included(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    groups: dict[tuple[str, ...], dict[str, Any]] = {}
    for record in records:
        if record.get("decision") != "include":
            continue
        model = normalize_space(record.get("normalized_model"))
        name = normalize_space(record.get("name"))
        package = normalize_space(record.get("package"))
        unit = normalize_space(record.get("unit"))
        qty = record.get("qty")
        merge_id = model if model else f"RAW:{name}:{record.get('source_file')}:{record.get('source_row')}"
        key = (record.get("category", ""), merge_id, package, unit)
        if key not in groups:
            groups[key] = {
                "category": record.get("category", ""),
                "model": model or name,
                "package": package,
                "raw_names": [],
                "qty": 0.0 if qty is not None else None,
                "unit": unit,
                "sources": [],
                "reasons": [],
            }
        group = groups[key]
        if qty is not None and group["qty"] is not None:
            group["qty"] += float(qty)
        else:
            group["qty"] = None
        group["raw_names"].append(normalize_space(record.get("name")))
        group["sources"].append(f"{record.get('source_file')}:{location_label(record)}")
        group["reasons"].append(record.get("reason", ""))

    output = []
    for group in groups.values():
        group["raw_name"] = "；".join(dict.fromkeys(value for value in group.pop("raw_names") if value))
        group["source"] = "；".join(dict.fromkeys(group.pop("sources")))
        group["decision_basis"] = "；".join(dict.fromkeys(value for value in group.pop("reasons") if value))
        if group["qty"] is not None and float(group["qty"]).is_integer():
            group["qty"] = int(group["qty"])
        output.append(group)
    return sorted(output, key=lambda item: (item["category"], item["model"], item["package"]))


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def write_markdown(path: Path, components: list[dict[str, Any]]):
    lines = [
        "| 类别 | 型号 / 规格 | 封装 / 安装形式 | 数量 | 单位 | 来源 |",
        "|---|---|---|---:|---|---|",
    ]
    for item in components:
        values = [item["category"], item["model"], item["package"], item["qty"], item["unit"], item["source"]]
        escaped = [str(value if value is not None else "").replace("|", "\\|") for value in values]
        lines.append("| " + " | ".join(escaped) + " |")
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")


def style_sheet(ws, column_widths: list[float]):
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_font = Font(name="宋体", size=10.5, bold=True, color="000000")
    data_font = Font(name="宋体", size=10.5, color="000000")
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            cell.border = border
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for idx, width in enumerate(column_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = width


def write_xlsx(path: Path, components: list[dict[str, Any]], review: list[dict[str, Any]]):
    workbook = openpyxl.Workbook()
    ws = workbook.active
    ws.title = "元器件清单"
    component_headers = ["序号", "类别", "型号 / 规格", "封装 / 安装形式", "原始品名", "数量", "单位", "来源", "判定依据"]
    ws.append(component_headers)
    for idx, item in enumerate(components, 1):
        ws.append([idx, item["category"], item["model"], item["package"], item["raw_name"], item["qty"], item["unit"], item["source"], item["decision_basis"]])
    style_sheet(ws, [8, 14, 32, 22, 34, 10, 8, 36, 24])

    review_ws = workbook.create_sheet("待复核")
    review_headers = ["来源文件", "位置", "原始文本", "名称", "型号 / 规格", "数量", "单位", "复核原因"]
    review_ws.append(review_headers)
    for item in review:
        review_ws.append([
            item.get("source_file", ""), location_label(item), item.get("raw_text", ""), item.get("name", ""),
            item.get("model", ""), item.get("qty"), item.get("unit", ""), item.get("reason", ""),
        ])
    style_sheet(review_ws, [24, 18, 60, 30, 32, 10, 8, 28])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def validate_xlsx(path: Path, component_count: int, review_count: int) -> dict[str, Any]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    component_rows = max(0, workbook["元器件清单"].max_row - 1)
    review_rows = max(0, workbook["待复核"].max_row - 1)
    workbook.close()
    checks = {
        "component_rows_match": component_rows == component_count,
        "review_rows_match": review_rows == review_count,
    }
    return {"component_rows": component_rows, "review_rows": review_rows, "checks": checks, "pass": all(checks.values())}


def export_outputs(input_path: Path, output_dir: Path, *, markdown_path: Path | None = None, output_xlsx: Path | None = None) -> dict[str, Any]:
    records, warnings, processed = extract_records(input_path)
    if not processed:
        raise ValueError("no supported PDF/XLSX/CSV invoice files found")

    included = [record for record in records if record.get("decision") == "include"]
    excluded = [record for record in records if record.get("decision") == "exclude"]
    review = [record for record in records if record.get("decision") == "review"]
    components = aggregate_included(records)

    output_dir.mkdir(parents=True, exist_ok=True)
    xlsx_path = output_xlsx or output_dir / "components.xlsx"
    components_csv = output_dir / "components.csv"
    review_csv = output_dir / "review.csv"
    normalized_csv = output_dir / "normalized_records.csv"
    report_path = output_dir / "validation_report.json"

    write_csv(
        components_csv,
        ["category", "model", "package", "raw_name", "qty", "unit", "source", "decision_basis"],
        components,
    )
    write_csv(
        review_csv,
        ["source_file", "source_type", "source_page", "source_sheet", "source_row", "raw_text", "name", "model", "qty", "unit", "reason"],
        review,
    )
    write_csv(
        normalized_csv,
        ["source_file", "source_type", "source_page", "source_sheet", "source_row", "raw_text", "name", "model", "qty", "unit", "decision", "reason", "category", "package", "normalized_model"],
        records,
    )
    write_xlsx(xlsx_path, components, review)
    if markdown_path:
        markdown_path.parent.mkdir(parents=True, exist_ok=True)
        write_markdown(markdown_path, components)

    xlsx_validation = validate_xlsx(xlsx_path, len(components), len(review))
    report = {
        "input": str(input_path),
        "processed_files": processed,
        "source_warnings": warnings,
        "counts": {
            "normalized_records": len(records),
            "included_records": len(included),
            "excluded_records": len(excluded),
            "review_records": len(review),
            "aggregated_components": len(components),
        },
        "outputs": {
            "xlsx": str(xlsx_path),
            "components_csv": str(components_csv),
            "review_csv": str(review_csv),
            "normalized_csv": str(normalized_csv),
            "markdown": str(markdown_path) if markdown_path else None,
        },
        "xlsx_validation": xlsx_validation,
    }
    report["pass"] = xlsx_validation["pass"] and not warnings
    report_path.write_text(json.dumps(report, ensure_ascii=False, indent=2), encoding="utf-8")
    return report


def is_soldering_component(name: str, model: str = "") -> bool:
    """Compatibility helper: only definite include returns True."""
    return classify_decision(name, model)[0] == "include"


def classify_component(name: str, model: str = ""):
    """Compatibility helper retained for older callers/tests."""
    return infer_category(name, model), clean_model(name, model), infer_package(name, model), normalize_space(f"{name} {model}")


def extract_soldering_components(invoices_dir, output_xlsx=None):
    """Compatibility wrapper around the new traceable workflow."""
    input_path = Path(invoices_dir).expanduser().resolve()
    output_path = Path(output_xlsx).expanduser().resolve() if output_xlsx else None
    output_dir = output_path.parent if output_path else input_path / "components_output"
    report = export_outputs(input_path, output_dir, output_xlsx=output_path)
    components = []
    with Path(report["outputs"]["components_csv"]).open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            row["qty"] = to_number(row.get("qty"))
            components.append(row)
    return components


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="采购资料焊接元器件提取与复核工具")
    parser.add_argument("--input", "--invoices-dir", dest="input_path", required=True, help="PDF/XLSX/CSV 文件或目录")
    parser.add_argument("--output-dir", help="输出目录；默认 <input>/components_output")
    parser.add_argument("--output-xlsx", help="兼容旧接口：指定 components.xlsx 路径")
    parser.add_argument("--markdown", help="可选 Markdown 清单输出路径")
    return parser


def main() -> int:
    args = build_parser().parse_args()
    input_path = Path(args.input_path).expanduser().resolve()
    output_xlsx = Path(args.output_xlsx).expanduser().resolve() if args.output_xlsx else None
    if args.output_dir:
        output_dir = Path(args.output_dir).expanduser().resolve()
    elif output_xlsx:
        output_dir = output_xlsx.parent
    elif input_path.is_dir():
        output_dir = input_path / "components_output"
    else:
        output_dir = input_path.parent / "components_output"
    markdown_path = Path(args.markdown).expanduser().resolve() if args.markdown else None

    try:
        report = export_outputs(input_path, output_dir, markdown_path=markdown_path, output_xlsx=output_xlsx)
    except Exception as exc:
        raise SystemExit(f"extraction failed: {exc}") from exc

    print(json.dumps(report, ensure_ascii=False, indent=2))
    return 0 if report["xlsx_validation"]["pass"] else 2


if __name__ == "__main__":
    raise SystemExit(main())
