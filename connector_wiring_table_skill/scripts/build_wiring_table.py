#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

AD_COLUMNS = ["BoardConnector", "BoardPin", "NetName"]
EXT_COLUMNS = ["SheetName", "CableEnd", "CablePin", "NetName", "TargetBoardConnector", "BoardPinHint", "CableConnectorModel", "Gender", "MatesTo"]
SIGNAL_COLUMNS = ["NetName", "SignalDefinition", "WireType", "ElectricalAttribute", "Include"]
OUT_HEADERS = ["序号", "连接点1代号", "节点号1", "", "连接点2代号", "节点号2", "线型", "信号定义", "电气属性/备注"]

class ValidationError(RuntimeError):
    pass

def read_csv(path: Path, required: List[str]) -> List[dict]:
    if not path.exists():
        raise ValidationError(f"缺少输入文件: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        missing = [c for c in required if c not in fields]
        if missing:
            raise ValidationError(f"{path.name} 缺少列: {', '.join(missing)}")
        rows = []
        for lineno, row in enumerate(reader, start=2):
            clean = {k: (v or "").strip() for k, v in row.items()}
            if any(clean.get(c, "") for c in required):
                clean["_line"] = lineno
                rows.append(clean)
        return rows

def truthy(value: str) -> bool:
    return value.strip().lower() in {"1", "true", "yes", "y", "是", "include"}

def build_normalized(ad_rows: List[dict], ext_rows: List[dict], signal_rows: List[dict]):
    errors: List[str] = []
    warnings: List[str] = []
    by_pin: Dict[Tuple[str, str], str] = {}
    by_connector_net: Dict[Tuple[str, str], List[str]] = defaultdict(list)
    for r in ad_rows:
        ref, pin, net = r["BoardConnector"], r["BoardPin"], r["NetName"]
        if not all([ref, pin, net]):
            errors.append(f"ad_pin_net.csv:{r['_line']} 存在空字段")
            continue
        pk = (ref, pin)
        if pk in by_pin and by_pin[pk] != net:
            errors.append(f"AD 引脚 {ref}.{pin} 同时属于 {by_pin[pk]} 与 {net}")
        by_pin[pk] = net
        key = (ref, net)
        if pin not in by_connector_net[key]:
            by_connector_net[key].append(pin)
    signals: Dict[str, dict] = {}
    for r in signal_rows:
        net = r["NetName"]
        if not net:
            errors.append(f"signal_catalog.csv:{r['_line']} NetName 为空")
            continue
        if net in signals:
            errors.append(f"signal_catalog.csv 中 NetName 重复: {net}")
        signals[net] = r
    cable_seen = set()
    board_seen = set()
    normalized = []
    sheet_order: List[str] = []
    connector_meta: Dict[str, dict] = {}
    for r in ext_rows:
        sheet, cable_end, cable_pin = r["SheetName"], r["CableEnd"], r["CablePin"]
        net, board_ref, hint = r["NetName"], r["TargetBoardConnector"], r["BoardPinHint"]
        if not all([sheet, cable_end, cable_pin, net, board_ref]):
            errors.append(f"external_pinout.csv:{r['_line']} 关键字段不能为空")
            continue
        if sheet not in sheet_order:
            sheet_order.append(sheet)
        ck = (cable_end, cable_pin)
        if ck in cable_seen:
            errors.append(f"外部连接器 pin 重复: {cable_end}.{cable_pin}")
        cable_seen.add(ck)
        sig = signals.get(net)
        if sig is None:
            errors.append(f"网络 {net} 未在 signal_catalog.csv 中定义")
            continue
        if not truthy(sig["Include"]):
            continue
        candidates = by_connector_net.get((board_ref, net), [])
        if hint:
            if hint not in candidates:
                errors.append(f"external_pinout.csv:{r['_line']} BoardPinHint={hint} 与 AD 中 {board_ref}/{net} 不一致，候选={candidates}")
                continue
            board_pin = hint
        elif len(candidates) == 1:
            board_pin = candidates[0]
        elif len(candidates) == 0:
            errors.append(f"external_pinout.csv:{r['_line']} 无法在 AD 中找到 {board_ref} 上网络 {net}")
            continue
        else:
            errors.append(f"external_pinout.csv:{r['_line']} {board_ref}/{net} 对应多个 pin {candidates}，必须填写 BoardPinHint")
            continue
        bk = (board_ref, board_pin)
        if bk in board_seen:
            errors.append(f"板端引脚被重复接出: {board_ref}.{board_pin}")
        board_seen.add(bk)
        model, gender, mates_to = r["CableConnectorModel"], r["Gender"], r["MatesTo"]
        meta = connector_meta.setdefault(cable_end, {"CableEnd": cable_end, "CableConnectorModel": model, "Gender": gender, "BoardConnector": board_ref, "MatesTo": mates_to})
        for k, v in [("CableConnectorModel", model), ("Gender", gender), ("BoardConnector", board_ref), ("MatesTo", mates_to)]:
            if meta[k] != v:
                errors.append(f"{cable_end} 的连接器元数据不一致: {k}={meta[k]} / {v}")
        normalized.append({"SheetName": sheet, "CableEnd": cable_end, "CablePin": cable_pin, "BoardConnector": board_ref, "BoardPin": board_pin, "NetName": net, "WireType": sig["WireType"], "SignalDefinition": sig["SignalDefinition"] or net, "ElectricalAttribute": sig["ElectricalAttribute"]})
    if errors:
        raise ValidationError("\n".join(errors))
    return normalized, sheet_order, list(connector_meta.values()), warnings

def write_normalized_csv(rows: List[dict], path: Path) -> None:
    fields = ["SheetName", "CableEnd", "CablePin", "BoardConnector", "BoardPin", "NetName", "WireType", "SignalDefinition", "ElectricalAttribute"]
    with path.open("w", encoding="utf-8-sig", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fields)
        w.writeheader()
        w.writerows(rows)

def style_sheet(ws):
    thin = Side(style="thin")
    fill = PatternFill("solid", fgColor="D9D9D9")
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)
            if cell.column != 4:
                cell.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for idx, width in enumerate([8, 16, 10, 3, 16, 10, 12, 24, 38], start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width
    ws.freeze_panes = "A2"

def write_xlsx(rows: List[dict], sheet_order: List[str], connector_meta: List[dict], path: Path):
    wb = Workbook()
    wb.remove(wb.active)
    for sheet in sheet_order:
        ws = wb.create_sheet(sheet[:31])
        ws.append(OUT_HEADERS)
        n = 0
        for r in rows:
            if r["SheetName"] != sheet:
                continue
            n += 1
            ws.append([n, r["CableEnd"], r["CablePin"], "", r["BoardConnector"], r["BoardPin"], r["WireType"], r["SignalDefinition"], r["ElectricalAttribute"]])
        ws.append([])
        ws.append(["说明", "节点号按连接器标准引脚号填写，不因公母视图反向而手动镜像。"])
        ws.append(["说明", "板端节点号由 AD pin/net 数据按 NetName 自动求得；存在歧义时必须使用 BoardPinHint。"])
        style_sheet(ws)
    ws = wb.create_sheet("连接器型号")
    ws.append(["线缆端编号", "连接器型号/规格", "公母/端接形式", "对接板端", "对接对象", "说明"])
    for m in connector_meta:
        ws.append([m["CableEnd"], m["CableConnectorModel"], m["Gender"], m["BoardConnector"], m["MatesTo"], ""])
    style_sheet(ws)
    wb.save(path)

def readback_rows(path: Path, sheet_order: List[str]):
    wb = load_workbook(path, data_only=False)
    got = []
    for sheet in sheet_order:
        ws = wb[sheet[:31]]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                break
            got.append((sheet, str(row[1]), str(row[2]), str(row[4]), str(row[5]), str(row[6] or ""), str(row[7] or ""), str(row[8] or "")))
    return got

def verify_readback(rows: List[dict], sheet_order: List[str], path: Path):
    expected = [(r["SheetName"], r["CableEnd"], r["CablePin"], r["BoardConnector"], r["BoardPin"], r["WireType"], r["SignalDefinition"], r["ElectricalAttribute"]) for r in rows]
    if readback_rows(path, sheet_order) != expected:
        raise ValidationError("生成后的 Excel 回读结果与 normalized_connections.csv 不一致")

def write_report(path: Path, rows: List[dict], warnings: List[str]):
    sheets = []
    for r in rows:
        if r["SheetName"] not in sheets:
            sheets.append(r["SheetName"])
    lines = ["# Wiring Table Validation Report", "", "## Result", "PASS", "", "## Summary", f"- Connections: {len(rows)}", f"- Wiring sheets: {len(sheets)}", "- Board pins resolved from AD net data: PASS", "- Excel read-back match: PASS", f"- Warnings: {len(warnings)}", "", "## Warnings"]
    lines.extend([f"- {w}" for w in warnings] or ["- None"])
    path.write_text("\n".join(lines) + "\n", encoding="utf-8")

def main():
    ap = argparse.ArgumentParser(description="Build and verify a connector wiring table.")
    ap.add_argument("input_dir", type=Path)
    ap.add_argument("output_dir", type=Path)
    args = ap.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    ad_rows = read_csv(args.input_dir / "ad_pin_net.csv", AD_COLUMNS)
    ext_rows = read_csv(args.input_dir / "external_pinout.csv", EXT_COLUMNS)
    signal_rows = read_csv(args.input_dir / "signal_catalog.csv", SIGNAL_COLUMNS)
    rows, sheet_order, connector_meta, warnings = build_normalized(ad_rows, ext_rows, signal_rows)
    write_normalized_csv(rows, args.output_dir / "normalized_connections.csv")
    xlsx = args.output_dir / "wiring_table.xlsx"
    write_xlsx(rows, sheet_order, connector_meta, xlsx)
    verify_readback(rows, sheet_order, xlsx)
    write_report(args.output_dir / "validation_report.md", rows, warnings)
    print(f"PASS: {len(rows)} connections -> {xlsx}")

if __name__ == "__main__":
    main()
