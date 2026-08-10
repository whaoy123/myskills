#!/usr/bin/env python3
from __future__ import annotations

import argparse
import csv
from collections import defaultdict
from pathlib import Path
from typing import Dict, List, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BOM_COLUMNS = ["Designator", "Model", "Footprint"]
RULE_COLUMNS = ["MatchField", "MatchValue", "DisplayPackage", "MountType", "PinsPerPart", "FixedPinsPerPart", "Include", "Notes"]
HEADERS = ["位号", "型号/规格", "封装", "数量", "贴片焊点", "直插焊点", "备注"]

class ValidationError(RuntimeError):
    pass

def read_csv(path: Path, required: List[str]) -> List[dict]:
    if not path.exists():
        raise ValidationError(f"缺少输入文件: {path}")
    with path.open("r", encoding="utf-8-sig", newline="") as f:
        reader = csv.DictReader(f)
        fields = reader.fieldnames or []
        missing = [c for c in required if c not in fields]
        if missing:
            raise ValidationError(f"{path.name} 缺少列: {', '.join(missing)}")
        rows = []
        for lineno, row in enumerate(reader, start=2):
            clean = {k: (v or "").strip() for k, v in row.items()}
            if any(clean.get(c, "") for c in required):
                clean["_line"] = lineno
                rows.append(clean)
        return rows

def truthy(v: str) -> bool:
    return v.strip().lower() in {"1", "true", "yes", "y", "是", "include"}

def to_int(v: str, what: str) -> int:
    try:
        n = int(v)
    except ValueError:
        raise ValidationError(f"{what} 必须为整数，实际为 {v!r}")
    if n < 0:
        raise ValidationError(f"{what} 不能为负数")
    return n

def build_rule_index(rule_rows: List[dict]):
    idx: Dict[Tuple[str, str], dict] = {}
    for r in rule_rows:
        field = r["MatchField"]
        if field not in {"Model", "Footprint"}:
            raise ValidationError(f"component_rules.csv:{r['_line']} MatchField 只能是 Model 或 Footprint")
        key = (field, r["MatchValue"])
        if not r["MatchValue"]:
            raise ValidationError(f"component_rules.csv:{r['_line']} MatchValue 为空")
        if key in idx:
            raise ValidationError(f"component_rules.csv 规则重复: {field}={r['MatchValue']}")
        mount = r["MountType"].upper()
        if mount not in {"SMD", "THT", "NONE"}:
            raise ValidationError(f"{field}={r['MatchValue']} 的 MountType 必须是 SMD/THT/NONE")
        r["_pins"] = to_int(r["PinsPerPart"], f"{field}={r['MatchValue']} PinsPerPart")
        r["_fixed"] = to_int(r["FixedPinsPerPart"], f"{field}={r['MatchValue']} FixedPinsPerPart")
        r["_mount"] = mount
        idx[key] = r
    return idx

def find_rule(row: dict, idx: Dict[Tuple[str, str], dict]) -> dict:
    for field in ("Model", "Footprint"):
        rule = idx.get((field, row[field]))
        if rule:
            return rule
    raise ValidationError(f"bom.csv:{row['_line']} {row['Designator']} ({row['Model']} / {row['Footprint']}) 没有匹配 component_rules.csv")

def normalize(bom_rows: List[dict], rule_rows: List[dict]):
    idx = build_rule_index(rule_rows)
    seen = set()
    groups = defaultdict(lambda: {"designators": []})
    excluded = 0
    for r in bom_rows:
        d = r["Designator"]
        if not d:
            raise ValidationError(f"bom.csv:{r['_line']} Designator 为空")
        if d in seen:
            raise ValidationError(f"BOM 位号重复: {d}")
        seen.add(d)
        rule = find_rule(r, idx)
        if not truthy(rule["Include"]) or rule["_mount"] == "NONE":
            excluded += 1
            continue
        key = (r["Model"], rule["DisplayPackage"], rule["_mount"], rule["_pins"], rule["_fixed"], rule["Notes"])
        g = groups[key]
        g.update({"model": r["Model"], "package": rule["DisplayPackage"] or r["Footprint"], "mount": rule["_mount"], "pins": rule["_pins"], "fixed": rule["_fixed"], "notes": rule["Notes"]})
        g["designators"].append(d)
    rows = []
    for g in groups.values():
        qty = len(g["designators"])
        per_part = g["pins"] + g["fixed"]
        smd = qty * per_part if g["mount"] == "SMD" else 0
        tht = qty * per_part if g["mount"] == "THT" else 0
        notes = g["notes"]
        if g["fixed"] and "固定脚" not in notes:
            extra = f"每只含{g['fixed']}个固定脚"
            notes = f"{notes}；{extra}" if notes else extra
        rows.append({"Designators": ", ".join(g["designators"]), "Model": g["model"], "Package": g["package"], "Quantity": qty, "SMDPoints": smd, "THTPoints": tht, "Notes": notes})
    rows.sort(key=lambda x: x["Designators"])
    return rows, excluded

def style(ws):
    thin = Side(style="thin")
    fill = PatternFill("solid", fgColor="D9D9D9")
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for row in ws.iter_rows(min_row=2):
        for c in row:
            c.alignment = Alignment(vertical="center", wrap_text=True)
            c.border = Border(top=thin, bottom=thin, left=thin, right=thin)
    for i, w in enumerate([32, 28, 24, 10, 12, 12, 32], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

def write_xlsx(rows: List[dict], path: Path):
    wb = Workbook()
    ws = wb.active
    ws.title = "焊接清单"
    ws.append(HEADERS)
    for r in rows:
        ws.append([r["Designators"], r["Model"], r["Package"], r["Quantity"], r["SMDPoints"], r["THTPoints"], r["Notes"]])
    qty, smd, tht = sum(r["Quantity"] for r in rows), sum(r["SMDPoints"] for r in rows), sum(r["THTPoints"] for r in rows)
    ws.append(["合计", "", "", qty, smd, tht, f"总焊点数：{smd + tht}"])
    style(ws)
    for col in (1, 4, 5, 6, 7):
        ws.cell(ws.max_row, col).font = Font(bold=True)
    wb.save(path)

def verify_readback(rows: List[dict], path: Path):
    wb = load_workbook(path, data_only=False)
    ws = wb["焊接清单"]
    got = []
    for values in ws.iter_rows(min_row=2, max_row=ws.max_row - 1, values_only=True):
        got.append({"Designators": str(values[0] or ""), "Model": str(values[1] or ""), "Package": str(values[2] or ""), "Quantity": int(values[3] or 0), "SMDPoints": int(values[4] or 0), "THTPoints": int(values[5] or 0), "Notes": str(values[6] or "")})
    if got != rows:
        raise ValidationError("生成后的 Excel 回读结果与脚本计算结果不一致")
    total = list(ws.iter_rows(min_row=ws.max_row, max_row=ws.max_row, values_only=True))[0]
    expected = (sum(r["Quantity"] for r in rows), sum(r["SMDPoints"] for r in rows), sum(r["THTPoints"] for r in rows))
    if (int(total[3]), int(total[4]), int(total[5])) != expected:
        raise ValidationError("Excel 合计行与脚本计算结果不一致")

def write_report(path: Path, rows: List[dict], excluded: int):
    qty, smd, tht = sum(r["Quantity"] for r in rows), sum(r["SMDPoints"] for r in rows), sum(r["THTPoints"] for r in rows)
    path.write_text(f"# Soldering Table Validation Report\n\n## Result\nPASS\n\n## Summary\n- Included component quantity: {qty}\n- Excluded BOM items: {excluded}\n- SMD solder joints: {smd}\n- THT solder joints: {tht}\n- Total solder joints: {smd + tht}\n- Excel read-back match: PASS\n", encoding="utf-8")

def main():
    ap = argparse.ArgumentParser(description="Build and verify a PCB soldering table.")
    ap.add_argument("input_dir", type=Path)
    ap.add_argument("output_dir", type=Path)
    args = ap.parse_args()
    args.output_dir.mkdir(parents=True, exist_ok=True)
    bom = read_csv(args.input_dir / "bom.csv", BOM_COLUMNS)
    rules = read_csv(args.input_dir / "component_rules.csv", RULE_COLUMNS)
    rows, excluded = normalize(bom, rules)
    out = args.output_dir / "soldering_table.xlsx"
    write_xlsx(rows, out)
    verify_readback(rows, out)
    write_report(args.output_dir / "validation_report.md", rows, excluded)
    print(f"PASS: {sum(r['Quantity'] for r in rows)} components -> {out}")

if __name__ == "__main__":
    main()
